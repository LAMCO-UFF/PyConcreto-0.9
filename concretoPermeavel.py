#Gerenciador de Janela - Concreto Permeável
import sys
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from openpyxl import Workbook

from materiais import *
from permOtim import *
from subsPerm import *
from subsPermot import *

class TrOt(QWidget):
    def __init__(self):
        super().__init__()
        #Esquerda
        self.p_c = QLineEdit()        
        self.p_w = QLineEdit()
        self.p_b = QLineEdit()        
        self.p_a = QLineEdit()
        self.ger_tr = QPushButton("Gerar traço")       
        self.ger_tr.setEnabled(False)
        #Direita
        self.m_c = QLineEdit()        
        self.m_w = QLineEdit()
        self.m_b = QLineEdit()        
        self.m_a = QLineEdit()
        self.save_tr = QPushButton("Salvar")       
        self.save_tr.setEnabled(False)
        #Inferior
        self.limpar = QPushButton("Limpar espaços")         
        self.sair = QPushButton("Sair")         #cria um botao para sair
        #Elementos gráficos
            #Boxes
        self.v0 = QVBoxLayout()
        self.h1 = QHBoxLayout()
        self.v2 = QVBoxLayout()
        self.v11 = QVBoxLayout()
        self.v12 = QVBoxLayout()
        self.v111 = QVBoxLayout()
        self.h112 = QHBoxLayout()
        self.v121 = QVBoxLayout()
        self.h122 = QHBoxLayout()
            #Esquerda
        self.placa_ro= QLabel("Massas específicas")
        self.placa_ro.setAlignment(Qt.AlignCenter)
        self.v111.addWidget(self.placa_ro)
        self.v111.addWidget(QLabel("Cimento (kg/m³)"))
        self.v111.addWidget(self.p_c)
        self.v111.addWidget(QLabel("Água (kg/m³)"))
        self.v111.addWidget(self.p_w)
        self.v111.addWidget(QLabel("Brita 0 (kg/m³)"))
        self.v111.addWidget(self.p_b)
        self.v111.addWidget(QLabel("Areia (kg/m³)"))
        self.v111.addWidget(self.p_a)
        self.h112.addWidget(self.ger_tr)
            #Direita
        self.placa_massas= QLabel("Traço otimizado (por m³ de concreto)")
        self.placa_massas.setAlignment(Qt.AlignCenter)
        self.v121.addWidget(self.placa_massas)
        self.v121.addWidget(QLabel("Cimento (kg)"))
        self.v121.addWidget(self.m_c)
        self.v121.addWidget(QLabel("Água (kg)"))
        self.v121.addWidget(self.m_w)
        self.v121.addWidget(QLabel("Brita 0 (kg)"))
        self.v121.addWidget(self.m_b)
        self.v121.addWidget(QLabel("Areia (kg)"))
        self.v121.addWidget(self.m_a)
        self.h122.addWidget(self.save_tr)
            #Inferior
        self.v2.addStretch()
        self.v2.addWidget(self.limpar)
        self.v2.addWidget(self.sair)
            #Implementação
        self.v0 = QVBoxLayout()         
        self.v0.addLayout(self.h1)
        self.h1.addLayout(self.v11)
        self.v11.addLayout(self.v111)
        self.v11.addLayout(self.h112)
        self.h1.addLayout(self.v12)
        self.v12.addLayout(self.v121)
        self.v12.addLayout(self.h122)
        self.v0.addLayout(self.v2)
        self.setLayout(self.v0)         

        #Sinais     (Conectam um elemento grafico a um metodo)
        self.p_c.textChanged[str].connect(self.tr_disable)
        self.p_w.textChanged[str].connect(self.tr_disable)
        self.p_b.textChanged[str].connect(self.tr_disable)
        self.p_a.textChanged[str].connect(self.tr_disable)
        self.ger_tr.clicked.connect(self.gerar_tr)
        
        self.m_c.textChanged[str].connect(self.save_disable)
        self.m_w.textChanged[str].connect(self.save_disable)
        self.m_b.textChanged[str].connect(self.save_disable)
        self.m_a.textChanged[str].connect(self.save_disable)
        self.save_tr.clicked.connect(self.salvar_tr)
        
        self.limpar.clicked.connect(self.limpar_itens)
        self.sair.clicked.connect(self.quit_application)

    #Eventos
    def tr_disable(self, s):         #verifica se pode acionar o botão
        if not self.p_c.text() or not self.p_w.text() or not self.p_b.text() or not self.p_a.text():
            self.ger_tr.setEnabled(False)
        else:
            self.ger_tr.setEnabled(True)

    def save_disable(self, s):         #verifica se pode acionar o botão
        if not self.m_c.text() or not self.m_w.text() or not self.m_b.text() or not self.m_a.text():
            self.save_tr.setEnabled(False)
        else:
            self.save_tr.setEnabled(True)

    def gerar_tr(self): 
        materiais=ConcPermOt()
        calculadora_trot=PermOt()
        
        materiais.mass_esp(float(self.p_c.text()), float(self.p_w.text()), float(self.p_b.text()), float(self.p_a.text()))
        fator=calculadora_trot.calc_vol(materiais.m_0, materiais.vet_ro)
        
        self.m_c.setText(str(round(materiais.m_0[0]/fator, 2)))
        self.m_w.setText(str(round(materiais.m_0[1]/fator, 2)))
        self.m_b.setText(str(round(materiais.m_0[2]/fator, 2)))
        self.m_a.setText(str(round(materiais.m_0[3]/fator, 2)))

    def salvar_tr(self):
        lista_materiais= ['Material', 'Cimento', 'Água', 'Brita 0', 'Areia']
        lista_tracos= ['Quantidade (kg/m³ de concreto)', self.m_c.text(), self.m_w.text(), self.m_b.text(), self.m_a.text()]
        arquivo_excel= Workbook()
        planilha1= arquivo_excel.active
        planilha1.title = 'Traço otimizado'
        for linha in range(5):
            planilha1.cell(row=linha+1, column=1, value=lista_materiais[linha])
            planilha1.cell(row=linha+1, column=2, value=lista_tracos[linha])
        arquivo_excel.save('traco_otimizado.xlsx')

    def limpar_itens(self):
        self.p_c.setText('')
        self.p_w.setText('')
        self.p_b.setText('')
        self.p_a.setText('')
        self.m_c.setText('')
        self.m_w.setText('')
        self.m_b.setText('')
        self.m_a.setText('')
        
    def quit_application(self):
        self.close()

class Subs(QWidget):
    def __init__(self):
        super().__init__()
        #Topo
        self.nome_ = QLineEdit()
        self.subs_ = QLineEdit()
        self.p_c = QLineEdit()        
        self.p_w = QLineEdit()
        self.p_b = QLineEdit()        
        self.p_m = QLineEdit()
        self.limpar = QPushButton("Limpar")
        #Meio-esquerda
        self.ger_tr1 = QPushButton("Calcular")       
        self.ger_tr1.setEnabled(False)
        self.m1_c = QLineEdit()        
        self.m1_w = QLineEdit()
        self.m1_b = QLineEdit()        
        self.m1_m = QLineEdit()
        self.limpar1 = QPushButton("Limpar")
        self.save_tr1 = QPushButton("Salvar")       
        self.save_tr1.setEnabled(False)
        #Meio-direita
        self.ger_tr2 = QPushButton("Calcular")       
        self.ger_tr2.setEnabled(False)
        self.m2_c = QLineEdit()        
        self.m2_w = QLineEdit()
        self.m2_b = QLineEdit()        
        self.m2_m = QLineEdit()
        self.limpar2 = QPushButton("Limpar")
        self.save_tr2 = QPushButton("Salvar")       
        self.save_tr2.setEnabled(False)
        #Inferior     
        self.sair = QPushButton("Sair")         #cria um botao para sair

        #Elementos gráficos
            #Boxes
        self.v0 = QVBoxLayout()
        self.v1 = QVBoxLayout()
        self.h2 = QHBoxLayout()
        self.h3 = QHBoxLayout()
        self.v21 = QVBoxLayout()
        self.v22 = QVBoxLayout()
            #Topo
        self.v1.addWidget(QLabel("Nome do material"))
        self.v1.addWidget(self.nome_)
        self.v1.addWidget(QLabel("Percentual de substituição (%)"))
        self.v1.addWidget(self.subs_)
        self.v1.addWidget(QLabel("Massa específica do material de substituição (kg/m³)"))
        self.v1.addWidget(self.p_m)
        self.v1.addWidget(QLabel("Massa específica do cimento (kg/m³)"))
        self.v1.addWidget(self.p_c)
        self.v1.addWidget(QLabel("Massa específica da água (kg/m³)"))
        self.v1.addWidget(self.p_w)
        self.v1.addWidget(QLabel("Massa específica da brita 0 (kg/m³)"))
        self.v1.addWidget(self.p_b)
        self.v1.addWidget(self.limpar)        
            #Meio-esquerda
        self.placa_c= QLabel("Substituição do cimento (por m³ de concreto)")
        self.placa_c.setAlignment(Qt.AlignCenter)
        self.v21.addWidget(self.placa_c)
        self.v21.addWidget(self.ger_tr1)
        self.v21.addWidget(QLabel("Cimento (kg)"))
        self.v21.addWidget(self.m1_c)
        self.v21.addWidget(QLabel("Água (kg)"))
        self.v21.addWidget(self.m1_w)
        self.v21.addWidget(QLabel("Brita 0 (kg)"))
        self.v21.addWidget(self.m1_b)
        self.v21.addWidget(QLabel("Material de substituição (kg)"))
        self.v21.addWidget(self.m1_m)
        self.v21.addWidget(self.limpar1)
        self.v21.addWidget(self.save_tr1)
            #Meio-direita
        self.placa_b= QLabel("Substituição da brita 0 (por m³ de concreto)")
        self.placa_b.setAlignment(Qt.AlignCenter)
        self.v22.addWidget(self.placa_b)
        self.v22.addWidget(self.ger_tr2)
        self.v22.addWidget(QLabel("Cimento (kg)"))
        self.v22.addWidget(self.m2_c)
        self.v22.addWidget(QLabel("Água (kg)"))
        self.v22.addWidget(self.m2_w)
        self.v22.addWidget(QLabel("Brita 0 (kg)"))
        self.v22.addWidget(self.m2_b)
        self.v22.addWidget(QLabel("Material de substituição (kg)"))
        self.v22.addWidget(self.m2_m)
        self.v22.addWidget(self.limpar2)
        self.v22.addWidget(self.save_tr2)
            #Inferior
        #self.v2.addStretch()
        self.h3.addWidget(self.sair)
            #Implementação         
        self.v0.addLayout(self.v1)
        self.v0.addLayout(self.h2)
        self.v0.addLayout(self.h3)
        self.h2.addLayout(self.v21)
        self.h2.addLayout(self.v22)
        self.setLayout(self.v0)         

        #Sinais     (Conectam um elemento grafico a um metodo)
        self.limpar.clicked.connect(self.limpar_0)

        self.p_c.textChanged[str].connect(self.tr_disable)
        self.p_w.textChanged[str].connect(self.tr_disable)
        self.p_b.textChanged[str].connect(self.tr_disable)
        self.p_m.textChanged[str].connect(self.tr_disable)
        
        self.ger_tr1.clicked.connect(self.gerar_tr1)
        self.limpar1.clicked.connect(self.limpar_1)
        self.m1_c.textChanged[str].connect(self.save1_disable)
        self.m1_w.textChanged[str].connect(self.save1_disable)
        self.m1_b.textChanged[str].connect(self.save1_disable)
        self.m1_m.textChanged[str].connect(self.save1_disable)
        self.save_tr1.clicked.connect(self.salvar_tr1)

        self.ger_tr2.clicked.connect(self.gerar_tr2)
        self.limpar2.clicked.connect(self.limpar_2)
        self.m2_c.textChanged[str].connect(self.save2_disable)
        self.m2_w.textChanged[str].connect(self.save2_disable)
        self.m2_b.textChanged[str].connect(self.save2_disable)
        self.m2_m.textChanged[str].connect(self.save2_disable)
        self.save_tr2.clicked.connect(self.salvar_tr2)
        
        self.sair.clicked.connect(self.quit_application)

    #Eventos
    def tr_disable(self, s):         #verifica se pode acionar o botão
        if not self.p_c.text() or not self.p_w.text() or not self.p_b.text() or not self.p_m.text():
            self.ger_tr1.setEnabled(False)
            self.ger_tr2.setEnabled(False)
        else:
            self.ger_tr1.setEnabled(True)
            self.ger_tr2.setEnabled(True)

    def save1_disable(self, s):         #verifica se pode acionar o botão
        if not self.m1_c.text() or not self.m1_w.text() or not self.m1_b.text() or not self.m1_m.text():
            self.save_tr1.setEnabled(False)
        else:
            self.save_tr1.setEnabled(True)

    def save2_disable(self, s):         #verifica se pode acionar o botão
        if not self.m2_c.text() or not self.m2_w.text() or not self.m2_b.text() or not self.m2_m.text():
            self.save_tr2.setEnabled(False)
        else:
            self.save_tr2.setEnabled(True)

    def gerar_tr1(self): 
        materiais1=ConcPermSub()
        calc_sub1=PermSub()
        
        materiais1.mass_esp(float(self.p_c.text()), float(self.p_w.text()), float(self.p_b.text()), float(self.p_m.text()))
        materiais1.teor_sub(float(self.subs_.text()), self.nome_.text())
        tal=calc_sub1.substituicao(materiais1.m_0[0], materiais1.sub)

        traco_sub=calc_sub1.calc_quant1(materiais1.m_0, tal, materiais1.vet_ro)
        
        self.m1_c.setText(str(round(traco_sub[0], 2)))
        self.m1_w.setText(str(round(traco_sub[1], 2)))
        self.m1_b.setText(str(round(traco_sub[2], 2)))
        self.m1_m.setText(str(round(traco_sub[3], 2)))

    def gerar_tr2(self): 
        materiais2=ConcPermSub()
        calc_sub2=PermSub()
        
        materiais2.mass_esp(float(self.p_c.text()), float(self.p_w.text()), float(self.p_b.text()), float(self.p_m.text()))
        materiais2.teor_sub(float(self.subs_.text()), self.nome_.text())
        tal=calc_sub2.substituicao(materiais2.m_0[2], materiais2.sub)

        traco_sub=calc_sub2.calc_quant2(materiais2.m_0, tal, materiais2.vet_ro)
        
        self.m2_c.setText(str(round(traco_sub[0], 2)))
        self.m2_w.setText(str(round(traco_sub[1], 2)))
        self.m2_b.setText(str(round(traco_sub[2], 2)))
        self.m2_m.setText(str(round(traco_sub[3], 2)))

    def salvar_tr1(self):
        material=self.nome_.text()
        lista_materiais= ['Material', 'Cimento', 'Água', 'Brita 0', material]
        lista_tracos= ['Quantidade (kg/m³ de concreto)', self.m1_c.text(), self.m1_w.text(), self.m1_b.text(), self.m1_m.text()]
        arquivo_excel= Workbook()
        planilha1= arquivo_excel.active
        planilha1.title = 'Traço de substituição'
        for linha in range(5):
            planilha1.cell(row=linha+1, column=1, value=lista_materiais[linha])
            planilha1.cell(row=linha+1, column=2, value=lista_tracos[linha])
        arquivo_excel.save('traco_substituicao_cimento.xlsx')

    def salvar_tr2(self):
        material=self.nome_.text()
        lista_materiais= ['Material', 'Cimento', 'Água', 'Brita 0', material]
        lista_tracos= ['Quantidade (kg/m³ de concreto)', self.m2_c.text(), self.m2_w.text(), self.m2_b.text(), self.m2_m.text()]
        arquivo_excel= Workbook()
        planilha1= arquivo_excel.active
        planilha1.title = 'Traço de substituição'
        for linha in range(5):
            planilha1.cell(row=linha+1, column=1, value=lista_materiais[linha])
            planilha1.cell(row=linha+1, column=2, value=lista_tracos[linha])
        arquivo_excel.save('traco_substituicao_brita.xlsx')

    def limpar_0(self):
        self.nome_.setText('')
        self.subs_.setText('')
        self.p_c.setText('')
        self.p_w.setText('')
        self.p_b.setText('')
        self.p_m.setText('')
        
    def limpar_1(self):
        self.m1_c.setText('')
        self.m1_w.setText('')
        self.m1_b.setText('')
        self.m1_m.setText('')

    def limpar_2(self):
        self.m2_c.setText('')
        self.m2_w.setText('')
        self.m2_b.setText('')
        self.m2_m.setText('')
        
    def quit_application(self):
        self.close()

class SubsOt(QWidget):
    def __init__(self):
        super().__init__()
        #Esquerda
        self.nome_ = QLineEdit()
        self.subs_ = QLineEdit()
        self.p_m = QLineEdit()
        self.p_c = QLineEdit()        
        self.p_w = QLineEdit()
        self.p_b = QLineEdit()        
        self.p_a = QLineEdit()
        self.limpar = QPushButton("Limpar")
        self.ger_tr1 = QPushButton("Calcular")       
        self.ger_tr1.setEnabled(False)
        #Direita
        self.m1_c = QLineEdit()        
        self.m1_w = QLineEdit()
        self.m1_b = QLineEdit()        
        self.m1_a = QLineEdit()
        self.m1_m = QLineEdit()
        self.save_tr1 = QPushButton("Salvar")       
        self.save_tr1.setEnabled(False)
        #Inferior     
        self.sair = QPushButton("Sair")         #cria um botao para sair

        #Elementos gráficos
            #Boxes
        self.v0 = QVBoxLayout()
        self.h1 = QHBoxLayout()
        self.v2 = QVBoxLayout()
        self.v11 = QVBoxLayout()
        self.v12 = QVBoxLayout()
            #Esquerda
        self.v11.addWidget(QLabel("Nome do material"))
        self.v11.addWidget(self.nome_)
        self.v11.addWidget(QLabel("Percentual de substituição (%)"))
        self.v11.addWidget(self.subs_)
        self.v11.addWidget(QLabel("Massa específica do material de substituição (kg/m³)"))
        self.v11.addWidget(self.p_m)
        self.v11.addWidget(QLabel("Massa específica do cimento (kg/m³)"))
        self.v11.addWidget(self.p_c)
        self.v11.addWidget(QLabel("Massa específica da água (kg/m³)"))
        self.v11.addWidget(self.p_w)
        self.v11.addWidget(QLabel("Massa específica da brita 0 (kg/m³)"))
        self.v11.addWidget(self.p_b)
        self.v11.addWidget(QLabel("Massa específica da areia (kg/m³)"))
        self.v11.addWidget(self.p_a)
        self.v11.addWidget(self.limpar)
        self.v11.addWidget(self.ger_tr1)
        self.v11.addStretch()
            #Direita
        self.v12.addWidget(QLabel("Cimento (kg/m³ de concreto)"))
        self.v12.addWidget(self.m1_c)
        self.v12.addWidget(QLabel("Água (kg/m³ de concreto)"))
        self.v12.addWidget(self.m1_w)
        self.v12.addWidget(QLabel("Brita 0 (kg/m³ de concreto)"))
        self.v12.addWidget(self.m1_b)
        self.v12.addWidget(QLabel("Areia (kg/m³ de concreto)"))
        self.v12.addWidget(self.m1_a)
        self.v12.addWidget(QLabel("Material de substituição (kg/m³ de concreto)"))
        self.v12.addWidget(self.m1_m)
        self.v12.addWidget(self.save_tr1)
        self.v12.addStretch()
            #Inferior
        self.v2.addStretch()
        self.v2.addWidget(self.sair)
            #Implementação         
        self.v0.addLayout(self.h1)
        self.v0.addLayout(self.v2)
        self.h1.addLayout(self.v11)
        self.h1.addLayout(self.v12)
        self.setLayout(self.v0)         

        #Sinais     (Conectam um elemento grafico a um metodo)
        self.p_c.textChanged[str].connect(self.tr_disable)
        self.p_w.textChanged[str].connect(self.tr_disable)
        self.p_b.textChanged[str].connect(self.tr_disable)
        self.p_a.textChanged[str].connect(self.tr_disable)
        self.p_m.textChanged[str].connect(self.tr_disable)        
        self.limpar.clicked.connect(self.limpar_0)
        self.ger_tr1.clicked.connect(self.gerar_tr1)
        
        self.m1_c.textChanged[str].connect(self.save1_disable)
        self.m1_w.textChanged[str].connect(self.save1_disable)
        self.m1_b.textChanged[str].connect(self.save1_disable)
        self.m1_a.textChanged[str].connect(self.save1_disable)
        self.m1_m.textChanged[str].connect(self.save1_disable)
        self.save_tr1.clicked.connect(self.salvar_tr1)
        
        self.sair.clicked.connect(self.quit_application)

    #Eventos
    def tr_disable(self, s):         #verifica se pode acionar o botão
        if not self.p_c.text() or not self.p_w.text() or not self.p_b.text() or not self.p_a.text() or not self.p_m.text():
            self.ger_tr1.setEnabled(False)
        else:
            self.ger_tr1.setEnabled(True)

    def save1_disable(self, s):         #verifica se pode acionar o botão
        if not self.m1_c.text() or not self.m1_w.text() or not self.m1_b.text() or not self.m1_m.text():
            self.save_tr1.setEnabled(False)
        else:
            self.save_tr1.setEnabled(True)

    def gerar_tr1(self): 
        materiais1=ConcPermSubOt()
        calc_sub1=PermSubOt()
        
        materiais1.mass_esp(float(self.p_c.text()), float(self.p_w.text()), float(self.p_b.text()), float(self.p_a.text()), float(self.p_m.text()))
        materiais1.teor_sub(float(self.subs_.text()), self.nome_.text())
        tal=calc_sub1.substituicao(materiais1.m_0[0], materiais1.sub)

        traco_sub=calc_sub1.calc_quant1(materiais1.m_0, tal, materiais1.vet_ro)
        
        self.m1_c.setText(str(round(traco_sub[0], 2)))
        self.m1_w.setText(str(round(traco_sub[1], 2)))
        self.m1_b.setText(str(round(traco_sub[2], 2)))
        self.m1_a.setText(str(round(traco_sub[3], 2)))
        self.m1_m.setText(str(round(traco_sub[4], 2)))

    def salvar_tr1(self):
        material=self.nome_.text()
        lista_materiais= ['Material', 'Cimento', 'Água', 'Brita 0', 'Areia', material]
        lista_tracos= ['Quantidade (kg/m³ de concreto)', self.m1_c.text(), self.m1_w.text(), self.m1_b.text(), self.m1_a.text(), self.m1_m.text()]
        arquivo_excel= Workbook()
        planilha1= arquivo_excel.active
        planilha1.title = 'Traço de substituição'
        for linha in range(6):
            planilha1.cell(row=linha+1, column=1, value=lista_materiais[linha])
            planilha1.cell(row=linha+1, column=2, value=lista_tracos[linha])
        arquivo_excel.save('traco_substituicao_otimizada_cimento.xlsx')

    def limpar_0(self):
        self.nome_.setText('')
        self.subs_.setText('')
        self.p_c.setText('')
        self.p_w.setText('')
        self.p_b.setText('')
        self.p_m.setText('')
        self.p_a.setText('')
        
    def quit_application(self):
        self.close()
